"""Improve PO matching between the DTC Retail tracker and the GL listing.

The tracker pulls actuals with SUMIFS against PA number + GL code + month.
Several formula bugs stop that from working, so GL journals never land on the
PA that raised them:

- Invoice-month header is `Jan` while Actuals stores `JAN` (case-sensitive).
- Local-currency SUMIFS were filled across, so the PA column walked into
  Type / Category / Description instead of staying on column J.
- One PA is split across several tracker rows (same PO + GL, different stores
  or tranches). Each row SUMIFS'd the *full* GL amount, so actuals doubled.
- GL `POR_Value` / LINDES arrive space-padded; tracker PAs do not.
- GB currency lookup pointed at tracker column K (Type) and a #REF! range.
- Actuals FX treated Danish kroner as `DDK`.

This script copies the workbook, fixes those formulas, adds normalised PO/GL
keys, and builds a PO_Match sheet so unmatched journals are visible.

Usage:
    python career/improve_budget_tracker.py
    python career/improve_budget_tracker.py --input "C:\\path\\to\\tracker.xlsx"
"""
from __future__ import annotations

import argparse
import re
import shutil
import zipfile
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import TableFormula

DEFAULT_INPUT = Path(
    r"C:\Users\akika\Downloads\DTC Retail 2026 Budget Tracker_Ruzana 1.xlsx"
)
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "DTC_Retail_2026_Budget_Tracker.xlsx"

TRACKER = "2026 DTC RETAIL TRACKER"
ACTUALS = "Actuals"
GL = "2026 GL listing"

# Tracker layout (1-based).
TRACKER_HEADER = 24
TRACKER_FIRST = 25
TRACKER_LAST = 2887
COL_STORE = 5          # E
COL_GL = 7             # G
COL_PA = 10            # J
COL_PA_VALUE = 20      # T
COL_INV_FIRST = 80     # CB invoice-currency Jan
COL_INV_LAST = 91      # CM Dec
COL_LOCAL_FIRST = 101  # CW local-currency Jan
COL_LOCAL_LAST = 112   # DH Dec
# Helpers sit just after ED (134): EE PO Key, EF GL Key, EG Weight, EH Match.
COL_PO_KEY = 135
COL_GL_KEY = 136
COL_WEIGHT = 137
COL_MATCH = 138

GL_HELPER_LAST = 6356  # query table is sized this far; formulas wait for a paste

C = {
    "dark": "1B3A4B",
    "sea": "2E6B7A",
    "red": "B42318",
    "warn": "B54708",
    "ok": "176F4A",
    "card": "F7F4EF",
    "line": "D6D0C7",
    "muted": "6B6560",
    "white": "FFFFFF",
    "red_bg": "FDE8E6",
    "ok_bg": "E7F5EE",
    "warn_bg": "FEF4E6",
    "blue": "1D4E89",
}


def _font(size=9, bold=False, color=C["dark"], name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    s = Side(style="thin", color=C["line"])
    return Border(left=s, right=s, top=s, bottom=s)


def _po_key_formula(cell: str) -> str:
    """Trim, drop spaces / non-breaking spaces, upper-case. Empty in → empty out."""
    return (
        f'IF({cell}="","",'
        f'UPPER(TRIM(SUBSTITUTE(SUBSTITUTE({cell}&""," ",""),CHAR(160),""))))'
    )


def restore_extlst(original: Path, modified: Path) -> int:
    """openpyxl drops x14 data-validation extensions; copy them back from source."""
    src = zipfile.ZipFile(original)
    dst = zipfile.ZipFile(modified)
    src_sheets = {n: src.read(n) for n in src.namelist() if n.startswith("xl/worksheets/sheet")}
    dst_names = dst.namelist()
    restored = 0
    tmp = modified.with_suffix(".xlsx.restoring")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for name in dst_names:
            data = dst.read(name)
            if name in src_sheets:
                orig = src_sheets[name].decode("utf-8", errors="replace")
                new = data.decode("utf-8", errors="replace")
                m = re.search(r"<extLst>.*?</extLst>", orig, re.DOTALL)
                if m and "<extLst>" not in new:
                    if new.endswith("</worksheet>"):
                        new = new[: -len("</worksheet>")] + m.group(0) + "</worksheet>"
                        restored += 1
                    data = new.encode("utf-8")
            out.writestr(name, data)
    src.close()
    dst.close()
    tmp.replace(modified)
    return restored


def fix_month_headers(ws) -> None:
    for col in (COL_INV_FIRST, COL_LOCAL_FIRST):
        if str(ws.cell(TRACKER_HEADER, col).value).strip() == "Jan":
            ws.cell(TRACKER_HEADER, col).value = "JAN"


def collect_tracker_rows(ws) -> list[int]:
    rows = []
    for r in range(TRACKER_FIRST, TRACKER_LAST + 1):
        pa = ws.cell(r, COL_PA).value
        inv = ws.cell(r, COL_INV_FIRST).value
        if pa or (isinstance(inv, str) and "SUMIFS" in inv):
            rows.append(r)
    return rows


def write_tracker_helpers_and_sumifs(ws, rows: list[int]) -> None:
    headers = [
        (COL_PO_KEY, "PO Key"),
        (COL_GL_KEY, "GL Key"),
        (COL_WEIGHT, "Alloc Weight"),
        (COL_MATCH, "Match vs GL"),
    ]
    for col, title in headers:
        c = ws.cell(TRACKER_HEADER, col)
        c.value, c.font, c.fill, c.alignment = (
            title, _font(8, bold=True, color=C["white"]),
            _fill(C["sea"]), _align("center"),
        )
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.cell(TRACKER_HEADER, COL_PO_KEY).comment = Comment(
        "Normalised PA/PO used to match GL journals. Do not type here.", "Finance"
    )
    ws.cell(TRACKER_HEADER, COL_WEIGHT).comment = Comment(
        "Share of GL actuals for this PO+GL code, based on PA value. "
        "Stops two tracker rows for the same PO from each pulling the full booking.",
        "Finance",
    )

    first, last = TRACKER_FIRST, TRACKER_LAST
    ee, ef, eh = (get_column_letter(c) for c in (COL_PO_KEY, COL_GL_KEY, COL_WEIGHT))
    t, j, g = (get_column_letter(c) for c in (COL_PA_VALUE, COL_PA, COL_GL))

    for r in rows:
        pa_cell = f"${j}{r}"
        gl_cell = f"${g}{r}"
        ws.cell(r, COL_PO_KEY).value = f"={_po_key_formula(pa_cell)}"
        ws.cell(r, COL_GL_KEY).value = f'=IF({gl_cell}="","",TRIM({gl_cell}&""))'
        # Allocation: split GL actuals across tracker rows that share PO + GL,
        # proportional to PA value. Equal split if PA values are blank.
        ws.cell(r, COL_WEIGHT).value = (
            f'=IF(${ee}{r}="","",'
            f'IF(SUMIFS(${t}${first}:${t}${last},${ee}${first}:${ee}${last},${ee}{r},'
            f"${ef}${first}:${ef}${last},${ef}{r})=0,"
            f"1/MAX(COUNTIFS(${ee}${first}:${ee}${last},${ee}{r},"
            f"${ef}${first}:${ef}${last},${ef}{r}),1),"
            f"${t}{r}/SUMIFS(${t}${first}:${t}${last},${ee}${first}:${ee}${last},${ee}{r},"
            f"${ef}${first}:${ef}${last},${ef}{r})))"
        )
        ws.cell(r, COL_MATCH).value = (
            f'=IF(${ee}{r}="","",'
            f'IF(COUNTIF(\'{GL}\'!$Q:$Q,${ee}{r})=0,"No GL booking",'
            f'IF(COUNTIFS(\'{GL}\'!$Q:$Q,${ee}{r},\'{GL}\'!$R:$R,${ef}{r})=0,'
            f'"GL booked to a different account/CC","Matched")))'
        )
        for col in (COL_PO_KEY, COL_GL_KEY, COL_WEIGHT, COL_MATCH):
            ws.cell(r, col).font = _font(8, color=C["muted"])
            ws.cell(r, col).fill = _fill(C["card"])

        for col in range(COL_INV_FIRST, COL_INV_LAST + 1):
            letter = get_column_letter(col)
            ws.cell(r, col).value = (
                f'=IF(${ee}{r}="","",'
                f"SUMIFS({ACTUALS}!$F:$F,{ACTUALS}!$N:$N,${ee}{r},"
                f"{ACTUALS}!$K:$K,UPPER({letter}$24),"
                f"{ACTUALS}!$B:$B,${ef}{r})*${eh}{r})"
            )
        for col in range(COL_LOCAL_FIRST, COL_LOCAL_LAST + 1):
            letter = get_column_letter(col)
            ws.cell(r, col).value = (
                f'=IF(${ee}{r}="","",'
                f"SUMIFS({ACTUALS}!$H:$H,{ACTUALS}!$N:$N,${ee}{r},"
                f"{ACTUALS}!$K:$K,UPPER({letter}$24),"
                f"{ACTUALS}!$B:$B,${ef}{r})*${eh}{r})"
            )

        # Sense-check was matching Category (L) to the PA number after a fill-right.
        ct = ws.cell(r, 98).value  # CT
        if isinstance(ct, str) and "$L:$L" in ct:
            ws.cell(r, 98).value = ct.replace("$L:$L", "$J:$J")


def fix_gl_listing(ws) -> int:
    """Normalised keys outside the query table so a refresh does not wipe them."""
    headers = [
        (17, "PO Key"),
        (18, "GL Key"),
        (19, "Match vs tracker"),
        (20, "Tracker PA value"),
        (21, "Tracker lines"),
    ]
    for col, title in headers:
        c = ws.cell(1, col)
        c.value, c.font, c.fill, c.alignment = (
            title, _font(9, bold=True, color=C["white"]),
            _fill(C["blue"]), _align("center", wrap=True),
        )
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.row_dimensions[1].height = 30

    # Fix the GB currency lookup: it pointed at Type (K) and a #REF! range.
    gb_currency = (
        f'=IFERROR(INDEX(\'{TRACKER}\'!$S${TRACKER_FIRST}:$S${TRACKER_LAST},'
        f"MATCH(Q{{r}},'{TRACKER}'!$EE${TRACKER_FIRST}:$EE${TRACKER_LAST},0)),\"\")"
    )
    match_key = '=IF(Q{r}="","",Q{r}&R{r})'
    month = (
        '=IFERROR(VLOOKUP(K{r},\'DATA VALIDATION\'!$D$3:$E$20,2,FALSE),"")'
    )
    amount = (
        '=D{r}+IF(H{r}="GB" ,E{r}*P{r},0)'
    )
    gb_fx = (
        '=IFERROR(IFS('
        'O{r}="USD",SUMPRODUCT((Inputs!$B$59:$B$63="USD")*(Inputs!$C$58:$N$58=N{r})*(Inputs!$C$59:$N$63)),'
        'O{r}="DKK",SUMPRODUCT((Inputs!$B$59:$B$63="DKK")*(Inputs!$C$58:$N$58=N{r})*(Inputs!$C$59:$N$63)),'
        'O{r}="SEK",SUMPRODUCT((Inputs!$B$59:$B$63="SEK")*(Inputs!$C$58:$N$58=N{r})*(Inputs!$C$59:$N$63)),'
        'O{r}="GBP",SUMPRODUCT((Inputs!$B$59:$B$63="GBP")*(Inputs!$C$58:$N$58=N{r})*(Inputs!$C$59:$N$63)),'
        'TRUE,SUMPRODUCT((Inputs!$B$59:$B$63="EUR")*(Inputs!$C$58:$N$58=N{r})*(Inputs!$C$59:$N$63))'
        '),0)'
    )

    filled = 0
    for r in range(2, GL_HELPER_LAST + 1):
        por = f"G{r}"
        lindes = f"J{r}"
        # Prefer POR_Value; fall back to LINDES (Aurora pads it to a fixed width).
        ws.cell(r, 17).value = (
            f'=IF(AND({por}="",{lindes}=""),"",'
            + _po_key_formula(f'IF({por}="",{lindes},{por})')
            + ")"
        )
        ws.cell(r, 18).value = f'=IF(A{r}="","",TRIM(A{r}&""))'
        ws.cell(r, 19).value = (
            f'=IF(Q{r}="","",'
            f'IF(COUNTIF(\'{TRACKER}\'!$EE:$EE,Q{r})=0,'
            f'IF(LEFT(Q{r},4)="2025","Prior-year PO — not on 2026 tracker","PO not on tracker"),'
            f'IF(COUNTIFS(\'{TRACKER}\'!$EE:$EE,Q{r},\'{TRACKER}\'!$EF:$EF,R{r})=0,'
            f'"PO on tracker — GL code not allocated","Matched")))'
        )
        ws.cell(r, 20).value = (
            f'=IF(Q{r}="","",SUMIF(\'{TRACKER}\'!$EE:$EE,Q{r},\'{TRACKER}\'!$T:$T))'
        )
        ws.cell(r, 21).value = (
            f'=IF(Q{r}="","",COUNTIF(\'{TRACKER}\'!$EE:$EE,Q{r}))'
        )
        ws.cell(r, 20).number_format = "#,##0.00"
        ws.cell(r, 12).value = match_key.format(r=r)       # L Match
        ws.cell(r, 13).value = amount.format(r=r)          # M AMOUNT invoiced
        ws.cell(r, 14).value = month.format(r=r)           # N Month invoiced
        ws.cell(r, 15).value = gb_currency.format(r=r)     # O GB currency
        ws.cell(r, 16).value = gb_fx.format(r=r)           # P GB_FX
        filled += 1

    # Keep the table's calculated-column templates in step with the cells,
    # otherwise a table resize reintroduces the broken #REF! lookup.
    tbl = ws.tables.get("_2026_GL_listing")
    if tbl is not None and not isinstance(tbl, str):
        templates = {
            "Match": "Q2&R2",
            "AMOUNT invoiced": 'D2+IF(H2="GB",E2*P2,0)',
            "Month invoiced": "IFERROR(VLOOKUP(K2,'DATA VALIDATION'!$D$3:$E$20,2,FALSE),\"\")",
            "GB_Invoice currency": (
                f"IFERROR(INDEX('{TRACKER}'!$S$25:$S$2887,"
                f"MATCH(Q2,'{TRACKER}'!$EE$25:$EE$2887,0)),\"\")"
            ),
        }
        for col in tbl.tableColumns:
            if col.name in templates:
                col.calculatedColumnFormula = TableFormula(attr_text=templates[col.name])
    return filled


def _cell_formula(value) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def fix_actuals(ws) -> int:
    """Point PA allocation at the normalised GL PO key; allow a manual override in P."""
    ws["P1"].value = "OPTIONAL"
    ws["P2"].value = "PA Override"
    ws["P2"].font = _font(9, bold=True, color=C["white"])
    ws["P2"].fill = _fill(C["sea"])
    ws["P2"].comment = Comment(
        "Leave blank to use the PO from the GL journal. Type a tracker PA here "
        "(e.g. POR10949) only when the journal PO is wrong or missing.",
        "Finance",
    )
    ws.column_dimensions["P"].width = 16

    notes = (
        f'=IF(N{{r}}="","",'
        f'IF(COUNTIFS(\'{TRACKER}\'!$EE:$EE,N{{r}},\'{TRACKER}\'!$EF:$EF,B{{r}})>0,"2026",'
        f'IF(COUNTIF(\'{TRACKER}\'!$EE:$EE,N{{r}})>0,"PO on tracker — GL code not allocated",'
        f'IF(LEFT(N{{r}},4)="2025","2025","Unmatched PO"))))'
    )

    n = 0
    for r in range(3, ws.max_row + 1):
        a = _cell_formula(ws.cell(r, 1).value)
        if not a or GL not in a:
            continue
        # Actuals row r reads GL row r-1 (A3 = GL!H2).
        gl_row = r - 1
        ws.cell(r, 2).value = f"='{GL}'!R{gl_row}"
        ws.cell(r, 14).value = (
            f'=UPPER(TRIM(IF(P{r}="",\'{GL}\'!Q{gl_row},P{r})&""))'
        )
        ws.cell(r, 15).value = notes.format(r=r)
        n += 1
    return n


def fix_actuals_fx(ws) -> int:
    fx_template = (
        '=IFERROR(IFS('
        'G{r}="USD",SUMPRODUCT((Inputs!$B$22:$B$37="USD")*(Inputs!$C$21:$N$21=K{r})*(Inputs!$C$22:$N$37)),'
        'G{r}="DKK",SUMPRODUCT((Inputs!$B$22:$B$37="DKK")*(Inputs!$C$21:$N$21=K{r})*(Inputs!$C$22:$N$37)),'
        'G{r}="SEK",SUMPRODUCT((Inputs!$B$22:$B$37="SEK")*(Inputs!$C$21:$N$21=K{r})*(Inputs!$C$22:$N$37)),'
        'G{r}="GBP",SUMPRODUCT((Inputs!$B$22:$B$37="GBP")*(Inputs!$C$21:$N$21=K{r})*(Inputs!$C$22:$N$37)),'
        'TRUE,SUMPRODUCT((Inputs!$B$22:$B$37="EUR")*(Inputs!$C$21:$N$21=K{r})*(Inputs!$C$22:$N$37))'
        '),0)'
    )
    n = 0
    for r in range(3, ws.max_row + 1):
        a = _cell_formula(ws.cell(r, 1).value)
        if not a or GL not in a:
            continue
        ws.cell(r, 10).value = fx_template.format(r=r)
        n += 1
    return n


def unique_pos(tracker_ws, gl_ws) -> list[str]:
    found: OrderedDict[str, None] = OrderedDict()
    for r in range(TRACKER_FIRST, TRACKER_LAST + 1):
        v = tracker_ws.cell(r, COL_PA).value
        if v and not str(v).startswith("="):
            found[str(v).strip().upper().replace(" ", "")] = None
    for r in range(2, gl_ws.max_row + 1):
        for col in (7, 10):  # POR_Value, LINDES
            v = gl_ws.cell(r, col).value
            if v and not str(v).startswith("="):
                found[str(v).strip().upper().replace(" ", "")] = None
    return [k for k in found if k]


def build_po_match(wb, pos: list[str]) -> None:
    if "PO_Match" in wb.sheetnames:
        del wb["PO_Match"]
    # Sit next to Actuals so the monthly review flow is Tracker → Actuals → match.
    idx = wb.sheetnames.index(ACTUALS) + 1 if ACTUALS in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("PO_Match", idx)
    ws.sheet_properties.tabColor = C["red"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"
    ws.page_setup.fitToPage = True

    ws.merge_cells("B1:I1")
    t = ws["B1"]
    t.value = "PO / PA match  —  tracker vs GL journals"
    t.font, t.alignment = _font(18, bold=True, color=C["dark"]), _align("left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("B2:I2")
    s = ws["B2"]
    s.value = (
        "Each unique PO from the tracker (column J) and the GL listing (POR_Value / LINDES). "
        "Status updates when you paste a new GL extract — no need to re-run Python unless new PAs appear."
    )
    s.font, s.alignment = _font(9, color=C["muted"]), _align("left", wrap=True)
    ws.row_dimensions[2].height = 32

    kpis = [
        (2, "Tracker PAs", f'=COUNTIF(B10:B{9+len(pos)+80},"Y")', C["blue"]),
        (4, "GL POs", f'=COUNTIF(E10:E{9+len(pos)+80},"Y")', C["sea"]),
        (6, "Matched", f'=COUNTIF(I10:I{9+len(pos)+80},"Matched")', C["ok"]),
        (8, "In GL, not on tracker", f'=COUNTIF(I10:I{9+len(pos)+80},"In GL — not on tracker")', C["red"]),
        (10, "On tracker, no GL", f'=COUNTIF(I10:I{9+len(pos)+80},"On tracker — no GL booking")', C["warn"]),
        (12, "Amount mismatch", f'=COUNTIF(I10:I{9+len(pos)+80},"Amount mismatch")', C["warn"]),
    ]
    for col, label, formula, color in kpis:
        lab = ws.cell(4, col)
        lab.value, lab.font, lab.fill, lab.alignment = (
            label, _font(8, bold=True, color=C["white"]), _fill(color), _align("center", wrap=True),
        )
        val = ws.cell(5, col)
        val.value, val.font, val.fill, val.alignment = (
            formula, _font(16, bold=True, color=color), _fill(C["card"]), _align("center"),
        )
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.row_dimensions[4].height = 28
        ws.row_dimensions[5].height = 32

    ws.merge_cells("B7:I7")
    how = ws["B7"]
    how.value = (
        "Matched = PO is on both sides and the GL account/CC exists on a tracker line. "
        "In GL — not on tracker = journal has a PO marketing has not entered (often a 2025xxxx placeholder). "
        "Add that PA on the tracker, or type the correct PA in Actuals column P. "
        "Amount mismatch = PO matches but booked GL ≠ sum of PA values — expected until invoices are fully in."
    )
    how.font, how.alignment = _font(8, color=C["muted"]), _align("left", wrap=True)
    ws.row_dimensions[7].height = 40

    headers = [
        (1, "PO Key", 16),
        (2, "In Tracker", 12),
        (3, "Tracker lines", 13),
        (4, "PA value", 14),
        (5, "In GL", 10),
        (6, "GL lines", 11),
        (7, "GL booked", 14),
        (8, "Variance (GL − PA)", 16),
        (9, "Status", 28),
    ]
    ws.row_dimensions[9].height = 22
    for col, title, width in headers:
        c = ws.cell(9, col)
        c.value, c.font, c.fill, c.alignment = (
            title, _font(9, bold=True, color=C["white"]),
            _fill(C["dark"]), _align("center", wrap=True),
        )
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = width

    spare = 80
    n_rows = len(pos) + spare
    for i in range(n_rows):
        r = 10 + i
        if i < len(pos):
            ws.cell(r, 1).value = pos[i]
        ws.cell(r, 2).value = f'=IF(A{r}="","",IF(COUNTIF(\'{TRACKER}\'!$EE:$EE,A{r})>0,"Y","N"))'
        ws.cell(r, 3).value = f'=IF(A{r}="","",COUNTIF(\'{TRACKER}\'!$EE:$EE,A{r}))'
        ws.cell(r, 4).value = f'=IF(A{r}="","",SUMIF(\'{TRACKER}\'!$EE:$EE,A{r},\'{TRACKER}\'!$T:$T))'
        ws.cell(r, 5).value = f'=IF(A{r}="","",IF(COUNTIF(\'{GL}\'!$Q:$Q,A{r})>0,"Y","N"))'
        ws.cell(r, 6).value = f'=IF(A{r}="","",COUNTIF(\'{GL}\'!$Q:$Q,A{r}))'
        ws.cell(r, 7).value = f'=IF(A{r}="","",SUMIF(\'{GL}\'!$Q:$Q,A{r},\'{GL}\'!$M:$M))'
        ws.cell(r, 8).value = f'=IF(A{r}="","",G{r}-D{r})'
        ws.cell(r, 9).value = (
            f'=IF(A{r}="","",'
            f'IF(AND(B{r}="N",E{r}="Y"),"In GL — not on tracker",'
            f'IF(AND(B{r}="Y",E{r}="N"),"On tracker — no GL booking",'
            f'IF(AND(B{r}="Y",E{r}="Y"),'
            f'IF(ABS(H{r})<1,"Matched","Amount mismatch"),""))))'
        )
        ws.cell(r, 4).number_format = '#,##0.00'
        ws.cell(r, 7).number_format = '#,##0.00'
        ws.cell(r, 8).number_format = '#,##0.00'
        for col in range(1, 10):
            c = ws.cell(r, col)
            c.font = _font(9)
            c.border = _border()
            c.alignment = _align("center" if col != 9 else "left")
            c.fill = _fill(C["card"] if i % 2 == 0 else C["white"])

    # Spare rows at the bottom so newly typed PAs are included without a rebuild.
    note_row = 10 + n_rows + 1
    ws.merge_cells(f"A{note_row}:I{note_row}")
    n = ws.cell(note_row, 1)
    n.value = (
        "Paste extra PO keys into column A of the blank rows above if a new PA is raised "
        "after this file was built. Filter Status for anything other than Matched."
    )
    n.font = _font(8, color=C["muted"])


def annotate_notes(ws) -> None:
    start = (ws.max_row or 46) + 2
    ws.cell(start, 2).value = "PO MATCHING (added)"
    ws.cell(start, 2).font = _font(12, bold=True, color=C["dark"])
    bullets = [
        "Tracker invoice actuals (CB:CM) and local actuals (CW:DH) now match GL journals on a normalised PO Key + 12-digit GL Key + month.",
        "Month headers are UPPERCASE so January (`Jan`) matches Actuals (`JAN`).",
        "Local-currency columns no longer walk off the PA column when filled right — they all stay on the PO Key.",
        "If the same PO + GL code appears on more than one tracker row, GL actuals are split by PA value (Alloc Weight in EG) instead of being counted twice.",
        "Filter `2026 GL listing` column S (Match vs tracker) or use the PO_Match tab. Placeholder POs like 20256044 will show as In GL — not on tracker until marketing enters them (or you override Actuals column P).",
        "Actuals column P is an optional PA override. Leave it blank to use the journal PO.",
    ]
    for i, text in enumerate(bullets):
        ws.cell(start + 1 + i, 2).value = text
        ws.cell(start + 1 + i, 2).alignment = _align("left", "top", wrap=True)
        ws.row_dimensions[start + 1 + i].height = 32
        ws.merge_cells(start_row=start + 1 + i, start_column=2, end_row=start + 1 + i, end_column=5)


def improve(src: Path, dest: Path) -> dict:
    if src.resolve() != dest.resolve():
        shutil.copy2(src, dest)

    wb = load_workbook(dest)
    tracker = wb[TRACKER]
    actuals = wb[ACTUALS]
    gl = wb[GL]

    fix_month_headers(tracker)
    rows = collect_tracker_rows(tracker)
    write_tracker_helpers_and_sumifs(tracker, rows)
    gl_helpers = fix_gl_listing(gl)
    actuals_n = fix_actuals(actuals)
    fix_actuals_fx(actuals)
    pos = unique_pos(tracker, gl)
    build_po_match(wb, pos)
    if "Notes" in wb.sheetnames:
        annotate_notes(wb["Notes"])

    wb.save(dest)
    wb.close()
    restored = restore_extlst(src, dest)
    return {
        "output": str(dest),
        "tracker_rows": len(rows),
        "gl_helper_rows": gl_helpers,
        "actuals_rows": actuals_n,
        "unique_pos": len(pos),
        "extlst_restored": restored,
        "sample_unmatched_gl": [p for p in pos if not p.startswith("POR")][:8],
    }


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Source tracker workbook")
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Where to write the improved copy")
    args = p.parse_args(argv)
    if not args.input.exists():
        raise SystemExit(f"Input not found: {args.input}")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    try:
        stats = improve(args.input, args.output)
    except PermissionError:
        alt = args.output.with_name(args.output.stem + "_matched.xlsx")
        print(f"{args.output.name} is open — writing {alt.name} instead")
        stats = improve(args.input, alt)
    print("Saved:", stats["output"])
    print(f"  Tracker rows rewired     {stats['tracker_rows']}")
    print(f"  GL helper rows           {stats['gl_helper_rows']}")
    print(f"  Actuals rows rewired     {stats['actuals_rows']}")
    print(f"  Unique POs on PO_Match   {stats['unique_pos']}")
    print(f"  Dropdown extensions kept {stats['extlst_restored']}")
    if stats["sample_unmatched_gl"]:
        print("  GL POs that are not POR-style (need a tracker line or Actuals P override):")
        print("   ", ", ".join(stats["sample_unmatched_gl"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
