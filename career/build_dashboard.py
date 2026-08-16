#!/usr/bin/env python3
"""
Supplier Management Dashboard — Excel workbook built from Aurora extracts.

Inputs:
  --input      AP_LG_NF open items (CSV)
  --reference  workbook holding the GL listing and vendor list (VENDMAST) tabs

Covers the four problem areas — late payment, invoice matching, invoice receipt
and direct debit control — plus the supplier reconciliation that sits under all
of them. KPI targets and SLAs come from "EMEA DTC Direct Debit Process" (7-day
reminder, 15-day escalation, B2B mandate migration).

The KPI registry below is the single source of truth: it drives both the
dashboard tiles and the KPI_Definitions sheet. Everything on the Dashboard is an
Excel formula over the source tables, so refreshing an extract in place
recalculates the workbook without re-running Python.

Usage:
  python career/build_dashboard.py --input "AP_LG_NF.CSV" --reference "GL+vendorlist.xlsx"
Output: career/supplier_dashboard.xlsx
"""
from __future__ import annotations

import argparse
import datetime
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from aurora_extracts import (
    GLLine,
    VendorLine,
    load_ap_log,
    load_gl_listing,
    load_vendor_list,
)

# Blank rows left inside each table so pasted data lands inside the table range
# and is picked up by the SUMIFS/COUNTIFS without anyone resizing anything.
GL_SPARE_ROWS = 250
VEND_SPARE_ROWS = 40
RECON_SPARE_ROWS = 40

# Reminder at 7 days, escalation 15 days after that (DD process doc, section 6).
SLA_REMINDER = 7
SLA_ESCALATE = SLA_REMINDER + 15

C = dict(
    dark="1A2332", accent="C45C26", sea="2F6F6A", warn="B8860B", ok="2D6A4F",
    muted="5A6678", red="C0392B", blue="2471A3", purple="6C3483", white="FFFFFF",
    line="E2D9CB", light="F8F5F0", card="FFFDF8", red_bg="FFF5F5", input_bg="FFFBEA",
)


def _fill(h):
    return PatternFill("solid", fgColor=h)


def _font(size=10, bold=False, color="1A2332", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin", color="E2D9CB"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, headers, colour, row=1, height=28):
    for j, (header, width) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = width
        c = ws.cell(row, j)
        c.value, c.font, c.fill, c.alignment = (
            header, _font(9, bold=True, color=C["white"]),
            _fill(colour), _align("center", wrap=True),
        )
    ws.row_dimensions[row].height = height


def _add_table(ws, name, n_cols, n_rows, style="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=f"A1:{get_column_letter(n_cols)}{1 + n_rows}")
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(tbl)


def _note(ws, row, text, span=8):
    c = ws.cell(row, 1)
    c.value = text
    c.font = _font(9, italic=True, color=C["muted"])
    c.alignment = _align("left", "top", wrap=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 42


# ─────────────────────────────────────────────────────────────────────
# KPI REGISTRY
# ─────────────────────────────────────────────────────────────────────
@dataclass
class KPI:
    ref: str
    name: str
    area: str
    definition: str
    formula: str
    fmt: str
    source: str
    owner: str
    target: str
    cadence: str
    colour: str = "blue"
    subtitle: str | None = None
    tile: bool = False


PAY, MATCH, RECEIPT = "Payment timeliness", "Invoice matching", "Invoice receipt"
DD, QUALITY, RECON, PL = ("Direct debit control", "Data quality",
                          "Supplier reconciliation", "P&L impact")

KPIS: list[KPI] = [
    KPI("K01", "Open AP (net)", PAY,
        "Total unpaid supplier balance in the AP log, net of credit notes.",
        "=SUM(AP_Data[Open Amount])", '"EUR "#,##0',
        "AP_Log", "AP", "Tracked, not targeted", "Weekly",
        "blue", '=COUNTA(AP_Data[Supplier Code])&" open items"', True),
    KPI("K02", "Overdue Value", PAY,
        "Open balance where the due date has passed.",
        '=SUMIF(AP_Data[Status],"Overdue",AP_Data[Open Amount])', '"EUR "#,##0',
        "AP_Log", "AP", "Reduce month on month", "Weekly",
        "red", '=COUNTIF(AP_Data[Status],"Overdue")&" invoices past due"', True),
    KPI("K03", "Overdue % of AP", PAY,
        "Overdue value divided by total open AP. The headline late-payment measure.",
        '=IFERROR(SUMIF(AP_Data[Status],"Overdue",AP_Data[Open Amount])'
        "/SUM(AP_Data[Open Amount]),0)", "0.0%",
        "AP_Log", "AP", "Below 10%", "Weekly",
        "warn", '="of EUR "&TEXT(SUM(AP_Data[Open Amount]),"#,##0")&" open"', True),
    KPI("K04", "Over 90 Days", PAY,
        "Open balance more than 90 days past due. Blocked items, not slow ones — "
        "a faster payment run will not clear them.",
        '=SUMIF(AP_Data[Aging Bucket],"90+",AP_Data[Open Amount])', '"EUR "#,##0',
        "AP_Log", "AP", "Zero", "Weekly",
        "red", '=COUNTIF(AP_Data[Aging Bucket],"90+")&" items stuck, not just late"', True),
    KPI("K05", "Weighted Avg Days Past Due", PAY,
        "Days past due weighted by open amount, so a large late invoice counts for "
        "more than a small one.",
        "=IFERROR(SUMPRODUCT(AP_Data[Days Past Due],AP_Data[Open Amount])"
        "/SUM(AP_Data[Open Amount]),0)", "0",
        "AP_Log", "AP", "Below 15 days", "Weekly",
        "warn", '="days, value weighted"', True),

    KPI("K06", "PO Coverage Rate", MATCH,
        "Share of AP-sourced GL lines carrying a POR reference. Rent and other "
        "contracted costs legitimately have none — they sit on the Contracted Cost "
        "Sheet — so read this alongside K07 rather than as a pass/fail.",
        '=IFERROR(COUNTIFS(GL_Data[Source],"A",GL_Data[PO Ref],"?*")'
        '/MAX(1,COUNTIF(GL_Data[Source],"A")),0)', "0.0%",
        "GL_Listing", "Retail / Business Owner", "Above 95% excluding rent", "Weekly",
        "sea", '="of AP-sourced GL lines carry a POR"', True),
    KPI("K07", "Match Exception Rate", MATCH,
        "Share of open AP lines carrying any exception flag.",
        '=IFERROR(COUNTIF(AP_Data[Exception Reason],"?*")'
        "/COUNTA(AP_Data[Supplier Code]),0)", "0.0%",
        "AP_Log", "AP", "Below 10%", "Weekly",
        "accent", '=COUNTIF(AP_Data[Exception Reason],"?*")&" lines flagged"', True),
    KPI("K08", "Value Without PO", MATCH,
        "Open AP value on invoices with no PO reference.",
        '=SUMIFS(AP_Data[Open Amount],AP_Data[PO Ref],"",AP_Data[Type],"IN")',
        '"EUR "#,##0',
        "AP_Log", "Retail / Business Owner", "Zero outside contracted costs",
        "Weekly", "accent", None, False),
    KPI("K09", "Unapplied Credits", MATCH,
        "Credit notes sitting unallocated. Until they net off, the gross invoice "
        "gets paid and the company overpays.",
        '=SUMIF(AP_Data[Type],"CR",AP_Data[Open Amount])', '"EUR "#,##0',
        "AP_Log", "AP", "Cleared within 30 days", "Monthly", "sea", None, False),

    KPI("K10", "Receipt SLA Breaches", RECEIPT,
        "Direct debit suppliers whose invoice is more than 7 calendar days past the "
        "expected submission date. The process doc requires a reminder at this point.",
        f'=COUNTIF(DD_Data[Days Past Expected],">{SLA_REMINDER}")', "0",
        "DD_Monitor", "AP", "Zero; reminder sent within 7 days", "Weekly",
        "warn", '="DD suppliers past the 7-day reminder point"', True),
    KPI("K11", "Receipt Escalations", RECEIPT,
        "Invoices still missing 15 calendar days after the reminder. The process doc "
        "requires escalation to the PO owner.",
        f'=COUNTIF(DD_Data[Days Past Expected],">{SLA_ESCALATE}")', "0",
        "DD_Monitor", "AP + Retail / Business Owner", "Zero", "Weekly",
        "red", None, False),
    KPI("K12", "Accrual Exposure", RECEIPT,
        "Expected but unreceived invoice value for DD suppliers, estimated from each "
        "supplier's average historical GL amount. A missing invoice is not in AP yet, "
        "so it has to be estimated rather than summed. Feeds the month-end accrual.",
        "=SUM(DD_Data[Accrual Estimate])", '"EUR "#,##0',
        "DD_Monitor", "Retail Finance", "Booked at month end", "Monthly",
        "purple", None, False),

    KPI("K13", "DD Without Active Mandate", DD,
        "Suppliers set to direct debit with no ACTIVE, treasury-signed mandate on "
        "file. Money can leave the account with no authorisation evidence.",
        '=COUNTIFS(Vend_Data[DD Supplier],"Y",Vend_Data[Mandate Status],"<>Active")',
        "0",
        "VENDMAST", "AP + Treasury", "Zero", "Monthly",
        "red", '="of "&COUNTIF(Vend_Data[DD Supplier],"Y")&" DD suppliers"', True),
    KPI("K14", "Core Scheme Share", DD,
        "Share of DD mandates on the Core scheme. Core mandates are not registered "
        "with HSBC, so anyone holding the bank details can collect.",
        '=IFERROR(COUNTIFS(Vend_Data[DD Supplier],"Y",Vend_Data[DD Scheme],"Core")'
        '/MAX(1,COUNTIF(Vend_Data[DD Supplier],"Y")),0)', "0.0%",
        "VENDMAST", "Treasury", "Migrate to B2B; currently circa 99% Core", "Quarterly",
        "warn", '="on Core — unregistered with the bank"', True),
    KPI("K15", "Mandate Documentation Complete", DD,
        "DD suppliers with an active mandate, signed date, treasury approval and "
        "SharePoint filing all present.",
        '=IFERROR(COUNTIFS(Vend_Data[DD Supplier],"Y",Vend_Data[Mandate Filed],"Y",'
        'Vend_Data[Mandate Status],"Active")'
        '/MAX(1,COUNTIF(Vend_Data[DD Supplier],"Y")),0)', "0.0%",
        "VENDMAST", "AP", "100%", "Quarterly", "sea", None, False),
    KPI("K16", "Manual Payments to DD Suppliers", DD,
        "Payments made by another method to a supplier set up for direct debit — "
        "either a failed collection or a duplicate payment risk.",
        '=SUMPRODUCT((AP_Data[Supplier Code]<>"")*(AP_Data[Pay Method]<>"DD")'
        '*COUNTIFS(Vend_Data[Supplier Code],AP_Data[Supplier Code]&"",'
        'Vend_Data[DD Supplier],"Y"))', "0",
        "AP_Log vs VENDMAST", "AP + Treasury", "Investigate every occurrence",
        "Weekly", "red", None, False),
    KPI("K23", "Direct Debit Exposure", DD,
        "Total open balance held by suppliers who can pull funds by direct debit. "
        "This is the amount collectable without a payment run approving it.",
        '=SUMIFS(Vend_Data[AP Balance],Vend_Data[DD Supplier],"Y")', '"EUR "#,##0',
        "VENDMAST", "Treasury", "Monitor", "Monthly",
        "purple", '=COUNTIF(Vend_Data[DD Supplier],"Y")&" suppliers can pull funds"',
        True),
    KPI("K24", "DD Suppliers With No Invoices", DD,
        "Direct debit suppliers with no invoice anywhere in the GL. Funds may be "
        "being collected with no supporting documentation — the exact gap the "
        "process document flags as having no systematic monitoring.",
        '=COUNTIF(DD_Data[Receipt Status],"No invoice activity")', "0",
        "DD_Monitor", "AP", "Zero", "Weekly",
        "red", '="DD suppliers with no invoice in the GL"', True),

    KPI("K17", "GL Codes Unusable", QUALITY,
        "AP lines whose GL account was rounded to scientific notation on export. The "
        "account family survives (613, 624, 628) but the specific account does not, "
        "so the line cannot be mapped to a P&L account.",
        '=SUMPRODUCT(--ISNUMBER(SEARCH("E+",AP_Data[GL Account]&"")))', "0",
        "AP_Log", "AP (fix the Aurora export)", "Zero", "Every extract",
        "red", '="lines cannot be mapped to P&L"', True),
    KPI("K18", "Suppliers Missing Payment Terms", QUALITY,
        "Vendor master records with no payment terms, which makes every invoice from "
        "them land overdue on arrival.",
        '=COUNTIFS(Vend_Data[Supplier Code],"?*",Vend_Data[Payment Terms (days)],"")',
        "0", "VENDMAST", "AP", "Zero", "Monthly", "warn", None, False),
    KPI("K19", "Top Supplier Concentration", QUALITY,
        "Largest single supplier as a share of open AP.",
        "=IFERROR(MAX(Supplier_Data[% of Open AP]),0)", "0.0%",
        "Suppliers", "Retail Finance", "Monitor above 25%", "Monthly",
        "sea", '="concentration across "&COUNTA(Supplier_Data[Supplier Code])'
               '&" suppliers"', True),

    KPI("K21", "Suppliers Not in Vendor Master", RECON,
        "Suppliers transacting in the AP log or GL with no vendor master record. "
        "Without one there is no payment method, no terms and no mandate — this is "
        "the root of 'payment or invoice hard to match to the supplier'.",
        '=COUNTIFS(Recon_Data[Supplier Code],"?*",'
        'Recon_Data[In Vendor Master],"No")', "0",
        "Reconciliation", "AP", "Zero", "Monthly",
        "red", '="trading suppliers with no master record"', True),
    KPI("K22", "AP vs Vendor Master Variance", RECON,
        "Absolute difference between the AP log open balance and the vendor master "
        "balance, for suppliers present in both. Compared only where both exist, "
        "because the extracts are samples taken at different times.",
        "=SUM(Recon_Data[Abs Variance])", '"EUR "#,##0',
        "Reconciliation", "AP", "Zero", "Monthly",
        "accent", '=COUNTIF(Recon_Data[Status],"Balance variance")'
                  '&" suppliers do not reconcile"', True),

    KPI("K20", "Spend by Flash Category", PL,
        "Posted supplier spend by Flash Category and period — the bridge from AP "
        "activity to the P&L. Accruals and their reversals are shown separately so "
        "period cost is not double counted.",
        "See P&L_View (SUMIFS over GL_Data by category, account and period)", "",
        "GL_Listing", "Retail Finance", "Against budget", "Monthly",
        "purple", None, False),
]

KPI_BY_REF = {k.ref: k for k in KPIS}

# Tiles are grouped by what the reader is asking, not by KPI number. Over-90-day
# balances sit with matching rather than timeliness because they are blocked
# items, and paying faster does not clear them.
BANDS: list[tuple[str, list[str]]] = [
    ("Payment timeliness", ["K01", "K02", "K03", "K05"]),
    ("Blocked items and invoice matching", ["K04", "K07", "K06", "K17"]),
    ("Direct debit control", ["K23", "K13", "K24", "K14"]),
    ("Invoice receipt and supplier reconciliation", ["K10", "K21", "K22", "K19"]),
]

TILE_KPIS = [k for k in KPIS if k.tile]
_banded = [ref for _, refs in BANDS for ref in refs]
assert sorted(_banded) == sorted(k.ref for k in TILE_KPIS), (
    "every tile KPI must appear in exactly one dashboard band"
)


# ─────────────────────────────────────────────────────────────────────
# SHEET: AP_LOG
# ─────────────────────────────────────────────────────────────────────
# Calculated columns use plain relative references ($H2) rather than the
# structured [@[Due Date]] shorthand. The shorthand is Excel UI sugar; written
# straight into the file it fails to parse and Excel strips the whole sheet's
# formulas on open. Templates take {r} = the row being written.
AP_COLS: list[tuple[str, int, str, str | None, str | None]] = [
    ("Supplier Code", 14, "center", None, None),
    ("Supplier Name", 20, "left", None, None),
    ("Type", 8, "center", None, None),
    ("LREF", 11, "center", None, None),
    ("Supplier Ref", 20, "left", None, None),
    ("PO Ref", 13, "center", None, None),
    ("Doc Date", 12, "center", "DD-MMM-YY", None),
    ("Due Date", 12, "center", "DD-MMM-YY", None),
    ("Period", 10, "center", None, None),
    ("Currency", 9, "center", None, None),
    ("Doc Amount", 13, "right", "#,##0.00", None),
    ("Open Amount", 13, "right", "#,##0.00", None),
    ("GL Account", 15, "center", None, None),
    ("Pay Method", 11, "center", None, None),
    ("Pay Cat", 8, "center", None, None),
    ("Ledger", 8, "center", None, None),
    ("Terms Days", 11, "center", "0",
     '=IF(OR($H{r}="",$G{r}=""),"",$H{r}-$G{r})'),
    # Always numeric: the weighted-average KPI multiplies this column, and a ""
    # here would poison the whole SUMPRODUCT with #VALUE!.
    ("Days Past Due", 12, "center", "0",
     '=IF($H{r}="",0,MAX(0,TODAY()-$H{r}))'),
    ("Status", 11, "center", None,
     '=IF($C{r}="CR","Credit",'
     'IF($H{r}="","No Due Date",'
     'IF($R{r}>0,"Overdue","Open")))'),
    ("Aging Bucket", 12, "center", None,
     '=IF(OR($S{r}="Credit",$S{r}="No Due Date"),"n/a",'
     'IF($R{r}=0,"Current",'
     'IF($R{r}<=30,"1-30",'
     'IF($R{r}<=60,"31-60",'
     'IF($R{r}<=90,"61-90","90+")))))'),
    ("In Vendor Master", 15, "center", None,
     '=IF($A{r}="","",'
     'IF(COUNTIF(Vend_Data[Supplier Code],$A{r})>0,"Yes","No"))'),
    # Priority order: the reason that blocks payment outranks the reason that
    # merely makes the line look odd.
    ("Exception Reason", 30, "left", None,
     '=IFS('
     '$C{r}="CR","Unapplied credit",'
     '$U{r}="No","Supplier not in vendor master",'
     'ISNUMBER(SEARCH("E+",$M{r}&"")),"GL account corrupted on export",'
     '$F{r}="","No PO reference",'
     '$Q{r}<0,"Due date before invoice date",'
     'AND($Q{r}<>"",$Q{r}<=1),"Payment terms 1 day or less",'
     '$K{r}<>$L{r},"Partial payment / amount mismatch",'
     'LEFT($F{r},3)<>"POR","Non-standard PO format",'
     '$R{r}>90,"Over 90 days past due",'
     'TRUE,"")'),
    # Ranking key for the dashboard exception queue. Absolute value so a large
    # credit surfaces too; the row fraction just breaks ties so MATCH is unique.
    ("Exc Sort Key", 12, "center", "0.000000",
     '=IF($V{r}="","",ABS($L{r})-ROW()/1000000)'),
]


def build_ap_log(ws, lines) -> None:
    ws.sheet_properties.tabColor = C["sea"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"
    _header_row(ws, [(h, w) for h, w, *_ in AP_COLS], C["sea"])

    for i, ln in enumerate(lines, 2):
        source = [
            ln.supplier_code, ln.supplier_name, ln.entry_type, ln.lref,
            ln.supplier_ref, ln.po_ref, ln.doc_date, ln.due_date, ln.period,
            ln.currency, ln.doc_amount, ln.open_amount, ln.gl_account,
            ln.pay_method, ln.pay_category, ln.ledger_code,
        ]
        ws.row_dimensions[i].height = 15
        for j, (_, _, al, fmt, formula) in enumerate(AP_COLS, 1):
            c = ws.cell(i, j)
            c.value = formula.format(r=i) if formula else source[j - 1]
            c.alignment = _align(al, "center")
            c.font = _font(9, color=C["dark"])
            if fmt:
                c.number_format = fmt
            if formula:
                c.fill = _fill(C["light"])

    _add_table(ws, "AP_Data", len(AP_COLS), len(lines))


# ─────────────────────────────────────────────────────────────────────
# SHEET: GL_LISTING
# ─────────────────────────────────────────────────────────────────────
GL_COLS: list[tuple[str, int, str | None, str | None]] = [
    ("Period", 10, None, None),
    ("Doc Date", 12, "DD-MMM-YY", None),
    ("GL Account", 15, None, None),
    ("Account Description", 30, None, None),
    ("Flash Category", 26, None, None),
    ("Channel", 10, None, None),
    ("Supplier Code", 13, None, None),
    ("Doc Ref", 12, None, None),
    ("PO Ref", 12, None, None),
    ("Description", 30, None, None),
    ("Amount", 13, "#,##0.00", None),
    ("Currency", 9, None, None),
    ("Source", 8, None, None),
    ("Doc Type", 9, None, None),
    ("Journal Type", 12, None, None),
    ("Ledger", 8, None, None),
    ("In Vendor Master", 15, None,
     '=IF($G{r}="","",'
     'IF(COUNTIF(Vend_Data[Supplier Code],$G{r})>0,"Yes","No"))'),
]


def build_gl_listing(ws, gl: list[GLLine]) -> None:
    ws.sheet_properties.tabColor = C["purple"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"
    _header_row(ws, [(h, w) for h, w, _, _ in GL_COLS], C["purple"])

    n_rows = len(gl) + GL_SPARE_ROWS
    for i in range(2, 2 + n_rows):
        idx = i - 2
        g = gl[idx] if idx < len(gl) else None
        source = [
            g.period, g.doc_date, g.gl_account, g.account_description,
            g.flash_category, g.channel, g.supplier_code, g.doc_ref, g.po_ref,
            g.description, g.amount, g.currency, g.source, g.doc_type,
            g.journal_type, g.ledger,
        ] if g else [None] * 16
        for j, (_, _, fmt, formula) in enumerate(GL_COLS, 1):
            c = ws.cell(i, j)
            c.value = formula.format(r=i) if formula else source[j - 1]
            c.font = _font(9)
            c.alignment = _align("left" if j in (4, 5, 10) else "center", "center")
            if fmt:
                c.number_format = fmt
            if g is None and not formula:
                c.fill = _fill(C["input_bg"])

    _add_table(ws, "GL_Data", len(GL_COLS), n_rows, "TableStyleLight9")
    _note(ws, 2 + n_rows + 2,
          "Aurora fields: PSTPER, DOCDT (CYYMMDD), ACCN08, PRLACC (supplier), DOCREF "
          "(= LREF on the AP log), LINDES (PO ref or free text), PSTAMT, TXSRCE "
          "(A = AP subledger, G = journal), Journal type (* = posted, Reversing = "
          "accrual). Export ACCN08 as TEXT — saving through CSV rounds a 12-digit "
          "account to 6.24E+11 and the line can no longer be mapped to a P&L account.",
          span=12)


# ─────────────────────────────────────────────────────────────────────
# SHEET: VENDMAST
# ─────────────────────────────────────────────────────────────────────
VEND_COLS: list[tuple[str, int, str | None, str | None]] = [
    ("Supplier Code", 14, None, None),
    ("Supplier Name", 18, None, None),
    ("Company", 10, None, None),
    ("Payment Method", 14, None, None),
    ("Currency", 9, None, None),
    ("AP Balance", 14, "#,##0.00", None),
    # Derived, so nobody has to keep a duplicate flag in step with PMTH05.
    ("DD Supplier", 12, None, '=IF($D{r}="DD","Y","N")'),
    ("DD Scheme", 11, None, None),
    ("Mandate Status", 14, None, None),
    ("Mandate Signed", 14, "DD-MMM-YY", None),
    ("Treasury Approved", 16, "DD-MMM-YY", None),
    ("Mandate Filed", 13, None, None),
    ("Expected Invoice Frequency", 22, None, None),
    ("Payment Terms (days)", 18, "0", None),
    ("Notes", 26, None, None),
]

FREQUENCIES = ["Weekly", "Fortnightly", "Monthly", "Quarterly", "Annual"]


def build_vendmast(ws, vendors: list[VendorLine]) -> int:
    ws.sheet_properties.tabColor = C["warn"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"
    _header_row(ws, [(h, w) for h, w, _, _ in VEND_COLS], C["warn"])

    n_rows = len(vendors) + VEND_SPARE_ROWS
    for i in range(2, 2 + n_rows):
        idx = i - 2
        v = vendors[idx] if idx < len(vendors) else None
        source = [v.supplier_code, v.supplier_name, v.company, v.pay_method,
                  v.currency, v.balance] if v else [None] * 6
        for j, (_, _, fmt, formula) in enumerate(VEND_COLS, 1):
            c = ws.cell(i, j)
            if formula:
                c.value = formula.format(r=i)
                c.fill = _fill(C["light"])
            elif j <= 6:
                c.value = source[j - 1]
                if v is None:
                    c.fill = _fill(C["input_bg"])
            else:
                c.fill = _fill(C["input_bg"])
            c.font = _font(9)
            c.alignment = _align("center" if j != 2 else "left", "center")
            if fmt:
                c.number_format = fmt

    for col, options in [
        ("H", '"Core,B2B"'),
        ("I", '"Active,Cancelled,Missing"'),
        ("L", '"Y,N"'),
        ("M", f'"{",".join(FREQUENCIES)}"'),
    ]:
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{1 + n_rows}")

    _add_table(ws, "Vend_Data", len(VEND_COLS), n_rows, "TableStyleLight10")
    _note(ws, 2 + n_rows + 2,
          "Columns A-F come straight from the Aurora vendor list (SUPN05, SNAM05, "
          "CONO05, PMTH05, CURN05, BLOG06). DD Supplier is derived from the payment "
          "method, so it never falls out of step. Columns H-O are not in Aurora — "
          "they come from the AP mandate repository and the contract, and they are "
          "what make the direct debit KPIs work.", span=12)
    return n_rows


# ─────────────────────────────────────────────────────────────────────
# SHEET: DD_MONITOR
# ─────────────────────────────────────────────────────────────────────
DD_COLS = [
    ("Supplier Code", 14, None),
    ("Supplier Name", 18, None),
    ("DD Supplier", 12, None),
    ("DD Scheme", 11, None),
    ("Mandate Status", 14, None),
    ("Mandate Complete", 16, None),
    ("Expected Frequency", 17, None),
    ("Expected Interval (days)", 19, "0"),
    ("Last Invoice in GL", 16, "DD-MMM-YY"),
    ("Days Since Last Invoice", 19, "0"),
    ("Days Past Expected", 16, "0"),
    ("Receipt Status", 19, None),
    ("Open AP", 13, "#,##0.00"),
    ("Avg Invoice Value", 15, "#,##0.00"),
    ("Accrual Estimate", 15, "#,##0.00"),
    ("Action Required", 40, None),
]

INTERVALS = [("Weekly", 7), ("Fortnightly", 14), ("Monthly", 30),
             ("Quarterly", 91), ("Annual", 365)]


def build_dd_monitor(ws, n_rows: int) -> None:
    ws.sheet_properties.tabColor = C["red"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"
    _header_row(ws, [(h, w) for h, w, _ in DD_COLS], C["red"], height=32)

    lookup = ",".join(f'"{f}",{d}' for f, d in INTERVALS)

    for i in range(2, 2 + n_rows):
        v = "VENDMAST!"
        formulas = [
            f'=IF({v}$A{i}="","",{v}$A{i})',
            f'=IF($A{i}="","",{v}$B{i})',
            f'=IF($A{i}="","",{v}$G{i})',
            f'=IF($A{i}="","",{v}$H{i})',
            f'=IF($A{i}="","",{v}$I{i})',
            f'=IF($A{i}="","",IF(AND({v}$I{i}="Active",{v}$J{i}<>"",'
            f'{v}$K{i}<>"",{v}$L{i}="Y"),"Yes","No"))',
            f'=IF($A{i}="","",{v}$M{i})',
            f'=IF($G{i}="","",IFERROR(SWITCH($G{i},{lookup}),""))',
            f'=IF($A{i}="","",IFERROR(MAXIFS(GL_Data[Doc Date],'
            f"GL_Data[Supplier Code],$A{i}),0))",
            f'=IF(OR($I{i}="",$I{i}=0),"",TODAY()-$I{i})',
            f'=IF(OR($J{i}="",$H{i}=""),"",$J{i}-$H{i})',
            # Receipt status is about receipt only; mandate gaps surface in K13
            # and in the Action column, so one does not mask the other.
            f'=IFS($A{i}="","",'
            f'$C{i}<>"Y","n/a (not DD)",'
            f'$I{i}=0,"No invoice activity",'
            f'$H{i}="","Frequency not set",'
            f"$K{i}>{SLA_ESCALATE},\"Escalate\","
            f"$K{i}>{SLA_REMINDER},\"Reminder due\","
            f'$K{i}>0,"Due",'
            f'TRUE,"On track")',
            f'=IF($A{i}="","",SUMIF(AP_Data[Supplier Code],$A{i},'
            f"AP_Data[Open Amount]))",
            # A missing invoice is not in AP yet, so the accrual has to be
            # estimated from what this supplier has historically billed.
            f'=IF($A{i}="","",IFERROR(AVERAGEIFS(GL_Data[Amount],'
            f'GL_Data[Supplier Code],$A{i},GL_Data[Source],"A"),0))',
            f'=IF(OR($L{i}="Reminder due",$L{i}="Escalate"),$N{i},0)',
            f'=IFS($A{i}="","",'
            f'AND($C{i}="Y",$F{i}="No"),"Obtain signed mandate + treasury sign-off",'
            f'$L{i}="No invoice activity","No invoice ever received — verify billing",'
            f'$L{i}="Frequency not set","Set expected invoice frequency in VENDMAST",'
            f'$L{i}="Escalate","Escalate to PO owner and contact supplier",'
            f'$L{i}="Reminder due","Send reminder to supplier",'
            f'TRUE,"")',
        ]
        ws.row_dimensions[i].height = 15
        for j, (formula, (_, _, fmt)) in enumerate(zip(formulas, DD_COLS), 1):
            c = ws.cell(i, j)
            c.value = formula
            c.font = _font(9)
            c.alignment = _align("center" if j not in (2, 16) else "left", "center")
            if fmt:
                c.number_format = fmt

    _add_table(ws, "DD_Data", len(DD_COLS), n_rows, "TableStyleMedium3")
    _note(ws, 2 + n_rows + 2,
          f"Thresholds follow the DD process document: reminder at {SLA_REMINDER} "
          f"calendar days past the expected invoice date, escalation to the PO owner "
          f"a further 15 days later ({SLA_ESCALATE} days total). Entirely "
          "formula-driven from VENDMAST and GL_Listing. Rows read 'Frequency not set' "
          "until the Expected Invoice Frequency column in VENDMAST is filled in — "
          "that is the one input the whole receipt control depends on.", span=12)


# ─────────────────────────────────────────────────────────────────────
# SHEET: RECONCILIATION
# ─────────────────────────────────────────────────────────────────────
REC_COLS = [
    ("Supplier Code", 14, "0"),
    ("Supplier Name", 18, None),
    ("In Vendor Master", 15, None),
    ("In AP Log", 11, None),
    ("In GL", 9, None),
    ("Payment Method", 14, None),
    ("Vendor Master Balance", 18, "#,##0.00"),
    ("AP Log Open", 14, "#,##0.00"),
    ("Variance", 13, "#,##0.00"),
    ("Abs Variance", 13, "#,##0.00"),
    ("GL Posted", 14, "#,##0.00"),
    ("Last GL Invoice", 14, "DD-MMM-YY"),
    ("Status", 22, None),
]


def build_reconciliation(ws, codes: list[str]) -> None:
    ws.sheet_properties.tabColor = C["accent"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"
    _header_row(ws, [(h, w) for h, w, _ in REC_COLS], C["accent"])

    n_rows = len(codes) + RECON_SPARE_ROWS
    for i in range(2, 2 + n_rows):
        idx = i - 2
        code = codes[idx] if idx < len(codes) else None
        ws.row_dimensions[i].height = 15
        values = [
            code,
            f'=IF($A{i}="","",IFERROR(INDEX(Vend_Data[Supplier Name],'
            f"MATCH($A{i},Vend_Data[Supplier Code],0)),\"\"))",
            f'=IF($A{i}="","",IF(COUNTIF(Vend_Data[Supplier Code],$A{i})>0,"Yes","No"))',
            f'=IF($A{i}="","",IF(COUNTIF(AP_Data[Supplier Code],$A{i})>0,"Yes","No"))',
            f'=IF($A{i}="","",IF(COUNTIF(GL_Data[Supplier Code],$A{i})>0,"Yes","No"))',
            f'=IF($A{i}="","",IFERROR(INDEX(Vend_Data[Payment Method],'
            f"MATCH($A{i},Vend_Data[Supplier Code],0)),\"\"))",
            f'=IF($A{i}="","",SUMIF(Vend_Data[Supplier Code],$A{i},'
            f"Vend_Data[AP Balance]))",
            f'=IF($A{i}="","",SUMIF(AP_Data[Supplier Code],$A{i},'
            f"AP_Data[Open Amount]))",
            # Only meaningful where the supplier appears in both extracts; the
            # AP log is a sample, so absence there is not a real variance.
            f'=IF($D{i}<>"Yes","",$H{i}-$G{i})',
            f'=IF($I{i}="",0,ABS($I{i}))',
            f'=IF($A{i}="","",SUMIF(GL_Data[Supplier Code],$A{i},GL_Data[Amount]))',
            f'=IF($A{i}="","",IFERROR(MAXIFS(GL_Data[Doc Date],'
            f"GL_Data[Supplier Code],$A{i}),0))",
            f'=IFS($A{i}="","",'
            f'$C{i}="No","Not in vendor master",'
            f'AND($D{i}="Yes",$J{i}>1),"Balance variance",'
            f'$D{i}="Yes","Reconciled",'
            f'TRUE,"Not in AP sample")',
        ]
        for j, (value, (_, _, fmt)) in enumerate(zip(values, REC_COLS), 1):
            c = ws.cell(i, j)
            c.value = value
            c.font = _font(9)
            c.alignment = _align("center" if j != 2 else "left", "center")
            if fmt and j != 1:
                c.number_format = fmt

    _add_table(ws, "Recon_Data", len(REC_COLS), n_rows, "TableStyleMedium6")
    _note(ws, 2 + n_rows + 2,
          "One row per supplier appearing in any of the three extracts. Variance is "
          "only calculated where the supplier is in the AP log, because that extract "
          "is a sample — a supplier missing from it is not a genuine difference. "
          '"Not in vendor master" is the serious one: a supplier being invoiced and '
          "posted with no master record has no agreed payment method, no terms and "
          "no mandate.", span=12)


# ─────────────────────────────────────────────────────────────────────
# SHEET: SUPPLIERS
# ─────────────────────────────────────────────────────────────────────
SUP_COLS = [
    ("Supplier Code", 14, "center", None),
    ("Supplier Name", 22, "left", None),
    ("Open Items", 11, "center", "0"),
    ("Open Amount", 15, "right", "#,##0.00"),
    ("Overdue Amount", 15, "right", "#,##0.00"),
    ("Over 90 Days", 14, "right", "#,##0.00"),
    ("Max Days Past Due", 16, "center", "0"),
    ("Exceptions", 11, "center", "0"),
    ("% of Open AP", 13, "center", "0.0%"),
]


def build_suppliers(ws, codes: list[str], names: dict[str, str]) -> None:
    ws.sheet_properties.tabColor = C["ok"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    _header_row(ws, [(h, w) for h, w, _, _ in SUP_COLS], C["ok"])

    for i, code in enumerate(codes, 2):
        ws.row_dimensions[i].height = 16
        values = [
            code, names[code],
            f"=COUNTIF(AP_Data[Supplier Code],$A{i})",
            f"=SUMIF(AP_Data[Supplier Code],$A{i},AP_Data[Open Amount])",
            f'=SUMIFS(AP_Data[Open Amount],AP_Data[Supplier Code],$A{i},'
            f'AP_Data[Status],"Overdue")',
            f'=SUMIFS(AP_Data[Open Amount],AP_Data[Supplier Code],$A{i},'
            f'AP_Data[Aging Bucket],"90+")',
            f"=IFERROR(MAXIFS(AP_Data[Days Past Due],AP_Data[Supplier Code],$A{i}),0)",
            f'=COUNTIFS(AP_Data[Supplier Code],$A{i},AP_Data[Exception Reason],"?*")',
            f"=IFERROR($D{i}/SUM(AP_Data[Open Amount]),0)",
        ]
        for j, (value, (_, _, al, fmt)) in enumerate(zip(values, SUP_COLS), 1):
            c = ws.cell(i, j)
            c.value, c.alignment, c.font = value, _align(al, "center"), _font(9)
            if fmt:
                c.number_format = fmt

    _add_table(ws, "Supplier_Data", len(SUP_COLS), len(codes), "TableStyleMedium4")


# ─────────────────────────────────────────────────────────────────────
# SHEET: P&L_VIEW
# ─────────────────────────────────────────────────────────────────────
def build_pl_view(ws, categories: list[str], accounts: list[tuple[str, str, str]],
                  year: int) -> None:
    ws.sheet_properties.tabColor = C["purple"]
    ws.sheet_view.showGridLines = False

    periods = [f"{year}-{m:02d}" for m in range(1, 13)]
    widths = [18, 30, 24] + [11] * 12 + [14]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    def header(row, first):
        cells = [first, "Description", "Flash Category"] + periods + ["Total"]
        for j, text in enumerate(cells, 1):
            c = ws.cell(row, j)
            c.value, c.font, c.fill, c.alignment = (
                text, _font(9, bold=True, color=C["white"]),
                _fill(C["purple"]), _align("center", wrap=True),
            )
        ws.row_dimensions[row].height = 26

    def money_row(row, key_formula_builder, bold=False):
        for m in range(12):
            col = 4 + m
            c = ws.cell(row, col)
            c.value = key_formula_builder(periods[m])
            c.number_format = "#,##0"
            c.font = _font(9, bold=bold)
        t = ws.cell(row, 16)
        t.value = f"=SUM(D{row}:O{row})"
        t.number_format = "#,##0"
        t.font = _font(9, bold=True)

    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = (
        f"Supplier spend by Flash Category and GL account — {year}. "
        "Driven by GL_Listing; accruals and reversals are split out below."
    )
    c.font, c.alignment = _font(12, bold=True), _align("left", "center")
    ws.row_dimensions[1].height = 24

    # ── Section 1: by Flash Category ─────────────────────────────────
    header(3, "Flash Category")
    row = 4
    cat_first = row
    for cat in categories:
        ws.cell(row, 1).value = cat
        ws.cell(row, 1).font = _font(9, bold=True)
        money_row(row, lambda p, cat=cat:
                  f'=SUMIFS(GL_Data[Amount],GL_Data[Flash Category],"{cat}",'
                  f'GL_Data[Period],"{p}")')
        row += 1
    for _ in range(4):
        ws.cell(row, 1).fill = _fill(C["input_bg"])
        money_row(row, lambda p, r=row:
                  f'=IF($A{r}="","",SUMIFS(GL_Data[Amount],'
                  f'GL_Data[Flash Category],$A{r},GL_Data[Period],"{p}"))')
        row += 1

    ws.cell(row, 1).value = "Total"
    ws.cell(row, 1).font = _font(10, bold=True)
    for col in range(4, 17):
        cl = get_column_letter(col)
        c = ws.cell(row, col)
        c.value = f"=SUM({cl}{cat_first}:{cl}{row - 1})"
        c.number_format = "#,##0"
        c.font = _font(10, bold=True)
        c.border = Border(top=Side(style="thin", color=C["dark"]))
    row += 1

    for label, criteria in [
        ("of which posted invoices", '"*"'),
        ("of which accruals & reversals", '"Reversing"'),
    ]:
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = _font(9, italic=True, color=C["muted"])
        money_row(row, lambda p, cr=criteria:
                  f"=SUMIFS(GL_Data[Amount],GL_Data[Journal Type],{cr},"
                  f'GL_Data[Period],"{p}")')
        row += 1

    # ── Section 2: by GL account ─────────────────────────────────────
    row += 2
    header(row, "GL Account")
    row += 1
    acc_first = row
    for account, description, category in accounts:
        ws.cell(row, 1).value = account
        ws.cell(row, 2).value = description
        ws.cell(row, 3).value = category
        for col in (1, 2, 3):
            ws.cell(row, col).font = _font(9)
            ws.cell(row, col).alignment = _align(
                "left" if col > 1 else "center", "center")
        money_row(row, lambda p, a=account:
                  f'=SUMIFS(GL_Data[Amount],GL_Data[GL Account],"{a}",'
                  f'GL_Data[Period],"{p}")')
        row += 1

    ws.cell(row, 1).value = "Total"
    ws.cell(row, 1).font = _font(10, bold=True)
    for col in range(4, 17):
        cl = get_column_letter(col)
        c = ws.cell(row, col)
        c.value = f"=SUM({cl}{acc_first}:{cl}{row - 1})"
        c.number_format = "#,##0"
        c.font = _font(10, bold=True)
        c.border = Border(top=Side(style="thin", color=C["dark"]))

    _note(ws, row + 2,
          "Accruals (Journal type 'Reversing') are posted in one period and reversed "
          "in the next, so a category total that looks volatile month to month is "
          "usually the accrual cycle rather than real spend. Compare the 'of which "
          "posted invoices' line for underlying cost. Accounts are 12 digits — any "
          "shown in scientific notation came through corrupted and will never match.",
          span=12)


# ─────────────────────────────────────────────────────────────────────
# SHEET: KPI_DEFINITIONS
# ─────────────────────────────────────────────────────────────────────
KPI_DEF_COLS = [
    ("Ref", 7, "center"), ("KPI", 26, "left"), ("Problem area", 18, "left"),
    ("Definition", 54, "left"), ("Excel formula", 46, "left"),
    ("Data source", 22, "left"), ("Owner", 20, "left"),
    ("Target / SLA", 26, "left"), ("Cadence", 12, "center"),
    ("On dashboard", 12, "center"),
]

AREA_COLOUR = {PAY: "red", MATCH: "accent", RECEIPT: "warn", DD: "purple",
               QUALITY: "sea", RECON: "blue", PL: "muted"}


def build_kpi_definitions(ws) -> None:
    ws.sheet_properties.tabColor = C["blue"]
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Supplier Management Dashboard — KPI definitions"
    c.font, c.fill, c.alignment = (
        _font(14, bold=True, color=C["white"]), _fill(C["dark"]),
        _align("left", "center"),
    )
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = (
        "Owner: Shan Jiang   ·   Stakeholders: DTC Retail Finance, AP   ·   "
        "Targets and SLAs per EMEA DTC Direct Debit Process"
    )
    c.font, c.fill, c.alignment = (
        _font(9, color="AAAAAA"), _fill(C["dark"]), _align("left", "center"),
    )
    ws.row_dimensions[2].height = 18

    _header_row(ws, [(h, w) for h, w, _ in KPI_DEF_COLS], C["blue"], row=3, height=24)
    ws.freeze_panes = "A4"

    for i, k in enumerate(sorted(KPIS, key=lambda x: x.ref), 4):
        ws.row_dimensions[i].height = 46
        values = [k.ref, k.name, k.area, k.definition, k.formula, k.source,
                  k.owner, k.target, k.cadence, "Yes" if k.tile else ""]
        for j, (value, (_, _, al)) in enumerate(zip(values, KPI_DEF_COLS), 1):
            c = ws.cell(i, j)
            c.value = value
            c.alignment = _align(al, "top", wrap=True)
            c.border = _border("thin", C["line"])
            if j == 1:
                c.font = _font(9, bold=True, color=C[AREA_COLOUR[k.area]])
            elif j == 5:
                c.font = Font(name="Consolas", size=8, color=C["muted"])
                # Documented for reference, so it must display as text rather
                # than being evaluated as a live formula.
                c.data_type = "s"
            else:
                c.font = _font(9, bold=(j == 2))


# ─────────────────────────────────────────────────────────────────────
# SHEET: DATA QUALITY
# ─────────────────────────────────────────────────────────────────────
DQ_ISSUES = [
    ("GL account corrupted on export",
     '=SUMPRODUCT(--ISNUMBER(SEARCH("E+",AP_Data[GL Account]&"")))',
     "Rounded to 3 significant figures. The account family survives (613, 624, 628) "
     "but the specific account does not, so the line cannot reach the P&L",
     "Yes — export ACCN08 as text"),
    ("Supplier not in vendor master",
     '=COUNTIFS(Recon_Data[Supplier Code],"?*",Recon_Data[In Vendor Master],"No")',
     "Trading with no master record: no agreed payment method, terms or mandate",
     "No — vendor master maintenance"),
    ("Vendor name truncated to 4 characters",
     '=SUMPRODUCT((Vend_Data[Supplier Code]<>"")'
     "*(LEN(Vend_Data[Supplier Name])<=4))",
     "SNAM05 arrives as 4 characters, so two different suppliers can both read "
     '"SCI " — the direct cause of payment and invoice being hard to match',
     "Yes — export the full supplier name"),
    ("AP line with no PO reference",
     '=COUNTIFS(AP_Data[PO Ref],"",AP_Data[Type],"IN")',
     "Nothing to three-way match against; invoice stalls indefinitely",
     "No — purchasing process"),
    ("GL line from AP with no PO reference",
     '=COUNTIFS(GL_Data[Source],"A",GL_Data[PO Ref],"")',
     "Mostly rent on the Contracted Cost Sheet, which is expected — review the rest",
     "No — review case by case"),
    ("Due date before invoice date",
     '=COUNTIF(AP_Data[Terms Days],"<0")',
     "Aging is meaningless; invoice is born overdue",
     "No — vendor master terms"),
    ("Payment terms 1 day or less",
     '=SUMPRODUCT((AP_Data[Terms Days]<>"")*(AP_Data[Terms Days]<=1)'
     "*(AP_Data[Terms Days]>=0))",
     "Vendor master terms missing, so every invoice lands overdue",
     "No — vendor master terms"),
    ("Unapplied credit note",
     '=COUNTIF(AP_Data[Type],"CR")',
     "Gross invoice gets paid and the credit never nets off",
     "No — AP process"),
    ("Missing due date",
     '=COUNTIF(AP_Data[Due Date],"")',
     "Excluded from aging entirely, so it is invisible to chasing",
     "Yes — PDUE should never be 00/00/00"),
    ("DD supplier without active mandate",
     '=COUNTIFS(Vend_Data[DD Supplier],"Y",Vend_Data[Mandate Status],"<>Active")',
     "Funds can be collected with no authorisation evidence on file",
     "No — AP mandate repository"),
    ("DD supplier with no expected frequency",
     '=COUNTIFS(Vend_Data[DD Supplier],"Y",'
     "Vend_Data[Expected Invoice Frequency],\"\")",
     "Without it there is no expected invoice date, so missing invoices go unnoticed",
     "No — set from contract and GL history"),
    ("Supplier missing payment terms",
     '=COUNTIFS(Vend_Data[Supplier Code],"?*",Vend_Data[Payment Terms (days)],"")',
     "Every invoice from this supplier arrives already overdue",
     "No — vendor master maintenance"),
]

DQ_COLS = [
    ("Export / data issue", 34, "left", None),
    ("Count", 8, "center", "0"),
    ("Why it matters", 58, "left", None),
    ("Fix at source?", 26, "left", None),
    ("Owner", 16, "left", None),
    ("Target date", 13, "center", "DD-MMM-YY"),
]


def build_data_quality(ws, serial_date_lines: int) -> None:
    ws.sheet_properties.tabColor = C["accent"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    _header_row(ws, [(h, w) for h, w, _, _ in DQ_COLS], C["accent"])

    rows = list(DQ_ISSUES)
    # Detected while parsing: once a serial reaches Excel it is a valid date,
    # so this one cannot be recomputed by a worksheet formula.
    rows.append(("Date exported as Excel serial", serial_date_lines,
                 "Mixed date formats in one column silently mis-bucket the aging",
                 "Yes — export DATE/PDUE as text"))

    for i, (issue, count, why, fix) in enumerate(rows, 2):
        ws.row_dimensions[i].height = 32
        for j, (value, (_, _, al, fmt)) in enumerate(
            zip([issue, count, why, fix, "", ""], DQ_COLS), 1
        ):
            c = ws.cell(i, j)
            c.value = value
            c.alignment = _align(al, "top", wrap=True)
            c.font = _font(9, bold=(j == 1))
            c.border = _border("thin", C["line"])
            if fmt:
                c.number_format = fmt
            if j in (5, 6):
                c.fill = _fill(C["input_bg"])

    _note(ws, len(rows) + 3,
          'Owner and target date are for you to fill in. Everything marked "Fix at '
          'source" has to be corrected in the Aurora export itself — no amount of '
          "downstream cleaning recovers a GL account already rounded to three "
          "significant figures, and saving a multi-tab workbook as .CSV silently "
          "drops every tab except the active one.", span=4)


# ─────────────────────────────────────────────────────────────────────
# SHEET: _CALC
# ─────────────────────────────────────────────────────────────────────
AGING_BUCKETS = ["Current", "1-30", "31-60", "61-90", "90+"]
EXCEPTION_REASONS = [
    "Unapplied credit", "Supplier not in vendor master",
    "GL account corrupted on export", "No PO reference",
    "Due date before invoice date", "Payment terms 1 day or less",
    "Partial payment / amount mismatch", "Non-standard PO format",
    "Over 90 days past due",
]
DD_STATUSES = ["On track", "Due", "Reminder due", "Escalate",
               "No invoice activity", "Frequency not set"]
EXC_SPILL_ROW = 40
EXC_DETAIL_ROWS = 20


def build_calc(ws, supplier_count: int, categories: list[str]) -> None:
    ws.sheet_state = "hidden"

    ws["A1"], ws["B1"] = "Bucket", "Open Amount"
    for i, bucket in enumerate(AGING_BUCKETS, 2):
        ws.cell(i, 1).value = bucket
        ws.cell(i, 2).value = f"=SUMIF(AP_Data[Aging Bucket],$A{i},AP_Data[Open Amount])"
        ws.cell(i, 2).number_format = "#,##0"

    last = supplier_count + 1
    ws["D1"], ws["E1"] = "Supplier", "Overdue Amount"
    for k in range(1, 6):
        i = k + 1
        ws.cell(i, 4).value = (
            f"=IFERROR(INDEX(Suppliers!$A$2:$A${last},"
            f"MATCH(LARGE(Suppliers!$E$2:$E${last},{k}),"
            f'Suppliers!$E$2:$E${last},0)),"")'
        )
        ws.cell(i, 5).value = f"=IFERROR(LARGE(Suppliers!$E$2:$E${last},{k}),0)"
        ws.cell(i, 5).number_format = "#,##0"

    ws["G1"], ws["H1"] = "Reason", "Lines"
    for i, reason in enumerate(EXCEPTION_REASONS, 2):
        ws.cell(i, 7).value = reason
        ws.cell(i, 8).value = f"=COUNTIF(AP_Data[Exception Reason],$G{i})"

    ws["J1"], ws["K1"] = "DD Status", "Suppliers"
    for i, status in enumerate(DD_STATUSES, 2):
        ws.cell(i, 10).value = status
        ws.cell(i, 11).value = f"=COUNTIF(DD_Data[Receipt Status],$J{i})"

    ws["M1"], ws["N1"] = "Flash Category", "Posted Spend"
    for i, cat in enumerate(categories, 2):
        ws.cell(i, 13).value = cat
        ws.cell(i, 14).value = (
            f"=SUMIFS(GL_Data[Amount],GL_Data[Flash Category],$M{i},"
            f'GL_Data[Journal Type],"*")'
        )
        ws.cell(i, 14).number_format = "#,##0"

    # Exception queue, largest exposure first. Built with LARGE + INDEX/MATCH on
    # the sort key rather than FILTER/SORT: dynamic arrays need spill metadata
    # that openpyxl cannot emit, and Excel discards the formulas on open.
    # Column P holds the nth largest key; columns A-G look the row up from it.
    queue_cols = [
        "AP_Data[Supplier Code]", "AP_Data[LREF]", "AP_Data[Supplier Ref]",
        "AP_Data[Due Date]", "AP_Data[Days Past Due]", "AP_Data[Open Amount]",
        "AP_Data[Exception Reason]",
    ]
    for n in range(EXC_DETAIL_ROWS):
        row = EXC_SPILL_ROW + n
        ws.cell(row, 16).value = f'=IFERROR(LARGE(AP_Data[Exc Sort Key],{n + 1}),"")'
        for j, column in enumerate(queue_cols, 1):
            ws.cell(row, j).value = (
                f'=IF($P{row}="","",IFERROR(INDEX({column},'
                f'MATCH($P{row},AP_Data[Exc Sort Key],0)),""))'
            )


# ─────────────────────────────────────────────────────────────────────
# SHEET: DASHBOARD
# ─────────────────────────────────────────────────────────────────────
DETAIL_COLS = [
    ("Supplier", "center", None), ("LREF", "center", None),
    ("Supplier Ref", "left", None), ("Due Date", "center", "DD-MMM-YY"),
    ("Days Past Due", "center", "0"), ("Open Amount", "right", "#,##0.00"),
    ("Exception Reason", "left", None),
]

BAND_ROWS = [5, 14, 23, 32]


def _kpi_block(ws, top_row: int, kpis: list[KPI], band: str) -> None:
    ws.row_dimensions[top_row].height = 16
    ws.merge_cells(f"B{top_row}:I{top_row}")
    c = ws.cell(top_row, 2)
    c.value, c.font, c.alignment = (
        band.upper(), _font(8, bold=True, color=C["muted"]), _align("left", "center"),
    )

    rows = list(range(top_row + 1, top_row + 8))
    for r, h in zip(rows, [10, 18, 34, 26, 10, 6, 6]):
        ws.row_dimensions[r].height = h

    for (cs, ce), k in zip([(2, 3), (4, 5), (6, 7), (8, 9)], kpis):
        colour = C[k.colour]
        csl, cel = get_column_letter(cs), get_column_letter(ce)
        for r in rows:
            for col in range(cs, ce + 1):
                cell = ws.cell(r, col)
                cell.fill = _fill(C["light"])
                cell.border = Border(
                    left=Side(style="thick", color=colour) if col == cs else Side(style="none"),
                    right=Side(style="thin", color=C["line"]) if col == ce else Side(style="none"),
                    top=Side(style="thin", color=C["line"]) if r == rows[0] else Side(style="none"),
                    bottom=Side(style="thin", color=C["line"]) if r == rows[-1] else Side(style="none"),
                )
            ws.merge_cells(f"{csl}{r}:{cel}{r}")

        c = ws.cell(top_row + 2, cs)
        c.value = f"{k.ref}  {k.name}"
        c.font, c.alignment = _font(8, bold=True, color="888888"), _align("center")

        c = ws.cell(top_row + 3, cs)
        c.value, c.number_format = k.formula, k.fmt
        c.font, c.alignment = _font(18, bold=True, color=colour), _align("center", "center")

        c = ws.cell(top_row + 4, cs)
        c.value = k.subtitle or f'="Target: {k.target}"'
        c.font = _font(8, italic=True, color=C["muted"])
        c.alignment = _align("center", "top", wrap=True)


def _section_header(ws, row: int, spans) -> None:
    ws.row_dimensions[row].height = 22
    for start, end, title in spans:
        ws.merge_cells(f"{start}{row}:{end}{row}")
        c = ws[f"{start}{row}"]
        c.value, c.font, c.alignment = title, _font(11, bold=True), _align("left", "center")


def _bar(ws_calc, cat_col, val_col, max_row, colour, horizontal, num_fmt):
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.style = 10
    chart.title = None
    chart.legend = None
    chart.y_axis.numFmt = num_fmt
    chart.width, chart.height = 14.5, 9
    chart.add_data(Reference(ws_calc, min_col=val_col, min_row=1, max_row=max_row),
                   titles_from_data=True)
    chart.set_categories(Reference(ws_calc, min_col=cat_col, min_row=2, max_row=max_row))
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = colour
        chart.series[0].graphicalProperties.line.solidFill = colour
    return chart


def build_dashboard(ws, ws_calc, source_name: str, n_categories: int) -> None:
    ws.sheet_properties.tabColor = C["accent"]
    ws.sheet_view.showGridLines = False

    for col, w in zip(range(1, 11), [2, 20, 16, 20, 16, 20, 16, 20, 16, 2]):
        ws.column_dimensions[get_column_letter(col)].width = w

    for r in range(1, 5):
        for col in range(1, 11):
            ws.cell(r, col).fill = _fill(C["dark"])
    for r, h in zip(range(1, 5), [8, 38, 18, 8]):
        ws.row_dimensions[r].height = h

    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value, c.font, c.alignment = (
        "Supplier Management Dashboard",
        _font(20, bold=True, color=C["white"]), _align("left", "center"),
    )

    ws.merge_cells("B3:I3")
    c = ws["B3"]
    c.value = (
        '="As of "&TEXT(TODAY(),"DD MMMM YYYY")'
        f'&"   ·   Sources: {source_name}, GL listing, vendor list'
        '  ·  KPI definitions on the KPI_Definitions tab"'
    )
    c.font, c.alignment = _font(9, color="AAAAAA"), _align("left", "center")

    for row, (title, refs) in zip(BAND_ROWS, BANDS):
        _kpi_block(ws, row, [KPI_BY_REF[r] for r in refs], title)

    _section_header(ws, 42, [("B", "E", "Aging Profile"),
                             ("F", "I", "Top 5 Overdue Suppliers")])
    ws.add_chart(_bar(ws_calc, 1, 2, 1 + len(AGING_BUCKETS), C["accent"],
                      False, '"EUR "#,##0'), "B43")
    ws.add_chart(_bar(ws_calc, 4, 5, 6, C["red"], True, '"EUR "#,##0'), "F43")

    _section_header(ws, 62, [("B", "E", "Exception Mix"),
                             ("F", "I", "Direct Debit Receipt Status")])
    ws.add_chart(_bar(ws_calc, 7, 8, 1 + len(EXCEPTION_REASONS), C["sea"],
                      True, "0"), "B63")
    ws.add_chart(_bar(ws_calc, 10, 11, 1 + len(DD_STATUSES), C["purple"],
                      True, "0"), "F63")

    _section_header(ws, 83, [("B", "E", "Posted Spend by Flash Category"),
                             ("F", "I", "How to read this")])
    ws.add_chart(_bar(ws_calc, 13, 14, 1 + max(n_categories, 1), C["blue"],
                      True, '"EUR "#,##0'), "B84")

    ws.merge_cells("F84:I100")
    c = ws["F84"]
    c.value = (
        "Aging answers whether we pay late. The 90+ bucket is different in kind — "
        "those items are blocked rather than slow, and paying faster will not clear "
        "them.\n\n"
        "Exception Mix answers why an item is blocked. Work it top down: the biggest "
        "bar is a process to fix, not an invoice to chase.\n\n"
        "Direct Debit Receipt Status is the control the DD process document asks for. "
        '"No invoice activity" means a supplier can pull funds but has never billed '
        "us — the highest risk row on the sheet. Statuses read \"Frequency not set\" "
        "until the Expected Invoice Frequency column in VENDMAST is filled in.\n\n"
        "Posted Spend excludes accruals and reversals, so it shows underlying cost "
        "rather than the accrual cycle.\n\n"
        "Reconciliation compares the AP log to the vendor master only where a "
        "supplier appears in both, because the extracts are samples."
    )
    c.font, c.alignment = _font(9, color=C["muted"]), _align("left", "top", wrap=True)

    _section_header(ws, 103, [("B", "I", "Exception Queue  (largest open amount first)")])
    hdr = 104
    ws.row_dimensions[hdr].height = 22
    for j, (header, _, _) in enumerate(DETAIL_COLS, 2):
        c = ws.cell(hdr, j)
        c.value, c.font, c.fill, c.alignment = (
            header, _font(9, bold=True, color=C["white"]),
            _fill(C["dark"]), _align("center"),
        )
        c.border = _border("thin", C["muted"])

    for i in range(EXC_DETAIL_ROWS):
        r = hdr + 1 + i
        calc_row = EXC_SPILL_ROW + i
        ws.row_dimensions[r].height = 16
        for j, (_, al, fmt) in enumerate(DETAIL_COLS, 2):
            calc_col = get_column_letter(j - 1)
            c = ws.cell(r, j)
            # _Calc already blanks unused queue rows, so no =0 guard here; that
            # would also hide a legitimate zero days-past-due or open amount.
            c.value = f'=IFERROR(_Calc!${calc_col}${calc_row},"")'
            c.fill = _fill(C["red_bg"] if i % 2 == 0 else C["card"])
            c.border = _border("thin", C["line"])
            c.alignment = _align(al, "center")
            c.font = _font(9, color=C["dark"])
            if fmt:
                c.number_format = fmt

    footer = hdr + EXC_DETAIL_ROWS + 2
    ws.merge_cells(f"B{footer}:I{footer}")
    c = ws.cell(footer, 2)
    c.value = (
        '=COUNTIF(AP_Data[Status],"Overdue")&" overdue  |  "'
        '&COUNTIF(AP_Data[Status],"Open")&" open  |  "'
        '&COUNTIF(AP_Data[Exception Reason],"?*")&" exceptions  |  "'
        '&COUNTIF(Recon_Data[Status],"Not in vendor master")&" suppliers missing a '
        'master record  |  Exception queue needs Excel 365 (FILTER/SORT)"'
    )
    c.font, c.alignment = _font(9, italic=True, color=C["muted"]), _align("left", "center")


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────
def main() -> int:
    here = Path(__file__).parent
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--input", type=Path, default=here / "ap_log.csv",
                    help="Aurora AP_LG_NF CSV export")
    ap.add_argument("--reference", type=Path, default=None,
                    help="workbook containing the GL listing and vendor list tabs")
    ap.add_argument("--gl-sheet", default="Sheet1")
    ap.add_argument("--vendor-sheet", default="Sheet2")
    ap.add_argument("--output", type=Path, default=here / "supplier_dashboard.xlsx")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"AP log not found: {args.input}")
        print('  python career/build_dashboard.py --input "C:\\path\\AP_LG_NF.CSV" '
              '--reference "C:\\path\\GL+vendorlist.xlsx"')
        return 1

    lines = load_ap_log(args.input)
    if not lines:
        print(f"No AP lines found in {args.input} — check the column headers.")
        return 1

    gl: list[GLLine] = []
    vendors: list[VendorLine] = []
    if args.reference:
        if not args.reference.exists():
            print(f"Reference workbook not found: {args.reference}")
            return 1
        gl = load_gl_listing(args.reference, args.gl_sheet)
        vendors = load_vendor_list(args.reference, args.vendor_sheet)

    names: dict[str, str] = {}
    for ln in lines:
        names.setdefault(ln.supplier_code, ln.supplier_name)
    codes = sorted(names)

    all_codes = sorted(
        {ln.supplier_code for ln in lines}
        | {g.supplier_code for g in gl if g.supplier_code}
        | {v.supplier_code for v in vendors}
    )

    categories = sorted({g.flash_category for g in gl if g.flash_category})
    accounts = sorted({
        (g.gl_account, g.account_description, g.flash_category) for g in gl
    })
    years = Counter(g.period[:4] for g in gl if len(g.period) == 7)
    if not years:
        years = Counter(ln.period[:4] for ln in lines if len(ln.period) == 7)
    year = int(years.most_common(1)[0][0]) if years else datetime.date.today().year

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    sheets = {
        name: wb.create_sheet(name) for name in
        ["KPI_Definitions", "AP_Log", "GL_Listing", "VENDMAST", "DD_Monitor",
         "Reconciliation", "Suppliers", "P&L_View", "Data_Quality", "_Calc"]
    }

    build_ap_log(sheets["AP_Log"], lines)
    build_gl_listing(sheets["GL_Listing"], gl)
    vend_rows = build_vendmast(sheets["VENDMAST"], vendors)
    build_dd_monitor(sheets["DD_Monitor"], vend_rows)
    build_reconciliation(sheets["Reconciliation"], all_codes)
    build_suppliers(sheets["Suppliers"], codes, names)
    build_pl_view(sheets["P&L_View"], categories, accounts, year)
    build_kpi_definitions(sheets["KPI_Definitions"])
    serial_dates = sum(1 for ln in lines if any("serial" in n for n in ln.quality_notes))
    build_data_quality(sheets["Data_Quality"], serial_dates)
    build_calc(sheets["_Calc"], len(codes), categories)
    build_dashboard(ws_dash, sheets["_Calc"], args.input.name, len(categories))

    wb.save(args.output)

    vend_codes = {v.supplier_code for v in vendors}
    gl_codes = {g.supplier_code for g in gl if g.supplier_code}
    orphans = sorted((gl_codes | set(codes)) - vend_codes) if vendors else []
    dd = [v for v in vendors if v.is_direct_debit]

    print(f"Saved: {args.output}")
    print(f"  AP lines   {len(lines):>4}   suppliers {len(codes)}")
    print(f"  GL lines   {len(gl):>4}   suppliers {len(gl_codes)}   "
          f"categories {len(categories)}   accounts {len(accounts)}")
    print(f"  Vendors    {len(vendors):>4}   direct debit {len(dd)}"
          f"   DD balance {sum(v.balance for v in dd):,.2f}")
    print(f"  KPIs       {len(KPIS):>4}   on dashboard {len(TILE_KPIS)}")
    if orphans:
        print(f"  Suppliers trading with no vendor master record: {len(orphans)}")
        print(f"    {', '.join(orphans)}")
    if dd:
        silent = [v.supplier_code for v in dd if v.supplier_code not in gl_codes]
        if silent:
            print(f"  DD suppliers with no GL invoice activity: {len(silent)} of "
                  f"{len(dd)}  ({', '.join(silent)})")
    flagged = sum(1 for ln in lines if ln.quality_notes)
    print(f"  AP lines with an export fault: {flagged}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
