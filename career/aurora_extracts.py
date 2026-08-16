"""Readers for the three Aurora extracts behind the supplier dashboard.

Aurora is an AS/400-era system and its exports show it: dates arrive as CYYMMDD
integers, periods as CYYPP, and code fields are space-padded to a fixed width.
Excel then adds its own damage on the way through — a 12-digit GL account saved
via CSV comes back as "6.24E+11" and can never be mapped to an account again.

Each loader returns plain dataclasses with the raw value preserved alongside the
parsed one, so the dashboard can report on export faults rather than hide them.
"""
from __future__ import annotations

import csv
import datetime
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# Excel's day zero, for AP log dates that were already converted to serials.
EXCEL_EPOCH = datetime.date(1899, 12, 30)

# A 12-digit account rounded to three significant figures still parses as a
# number, so the scientific-notation text is the only reliable tell.
GL_CORRUPT_HINT = "E+"

PO_PATTERN = re.compile(r"^POR\d+", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────────────
# Shared field parsers
# ─────────────────────────────────────────────────────────────────────
def parse_cyymmdd(raw) -> datetime.date | None:
    """Aurora date: 1260204 -> 2026-02-04. Leading digit is the century flag."""
    if raw in (None, "", 0):
        return None
    s = str(raw).strip()
    if not s.isdigit() or len(s) != 7:
        return None
    century, yy, mm, dd = int(s[0]), int(s[1:3]), int(s[3:5]), int(s[5:7])
    if not (1 <= mm <= 12 and 1 <= dd <= 31):
        return None
    try:
        return datetime.date(1900 + century * 100 + yy, mm, dd)
    except ValueError:
        return None


def parse_cyypp(raw) -> str:
    """Aurora period: 12601 -> '2026-01'. AP log uses a 4-digit 2606 variant."""
    if raw in (None, ""):
        return ""
    s = str(raw).strip()
    if len(s) == 5 and s.isdigit():
        return f"{1900 + int(s[0]) * 100 + int(s[1:3])}-{s[3:5]}"
    if len(s) == 4 and s.isdigit():
        return f"20{s[:2]}-{s[2:]}"
    return s


def _clean(value) -> str:
    return str(value).strip() if value is not None else ""


def _amount(raw) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    return float(str(raw).strip().replace(",", "") or 0)


def _find_header_row(ws, marker: str, limit: int = 20) -> int:
    """Aurora exports carry title and control-total rows above the headers."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True), 1):
        if any(_clean(v).upper() == marker.upper() for v in row):
            return i
    raise ValueError(f"Could not find a header row containing {marker!r}")


def _column_map(ws, header_row: int) -> dict[str, int]:
    return {
        _clean(c.value): c.column
        for c in ws[header_row]
        if _clean(c.value)
    }


# ─────────────────────────────────────────────────────────────────────
# AP log — AP_LG_NF open items
# ─────────────────────────────────────────────────────────────────────
@dataclass
class APLine:
    supplier_code: str
    supplier_name: str
    entry_type: str
    lref: str
    supplier_ref: str
    po_ref: str
    doc_date: datetime.date | None
    due_date: datetime.date | None
    period: str
    currency: str
    doc_amount: float
    open_amount: float
    gl_account: str
    pay_method: str
    pay_category: str
    ledger_code: str
    quality_notes: list[str] = field(default_factory=list)


def _ap_date(raw: str, notes: list[str], label: str) -> datetime.date | None:
    """AP log emits dd/mm/yy, a bare Excel serial, or 00/00/00 for 'none'."""
    raw = _clean(raw)
    if not raw or raw == "00/00/00":
        return None
    if raw.isdigit():
        notes.append(f"{label} exported as Excel serial number")
        return EXCEL_EPOCH + datetime.timedelta(days=int(raw))
    try:
        return datetime.datetime.strptime(raw, "%d/%m/%y").date()
    except ValueError:
        notes.append(f"{label} unparseable ({raw!r})")
        return None


def load_ap_log(path: Path) -> list[APLine]:
    lines: list[APLine] = []
    with Path(path).open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if not _clean(row.get("SUPN15")):
                continue
            notes: list[str] = []
            gl = _clean(row.get("GLAC17"))
            if GL_CORRUPT_HINT in gl.upper():
                notes.append("GL account truncated to scientific notation on export")
            doc_amount = _amount(row.get("BTMT17"))
            open_amount = _amount(row.get("PTMT17"))
            if doc_amount != open_amount:
                notes.append("Document amount differs from open amount")
            lines.append(APLine(
                supplier_code=_clean(row.get("SUPN15")),
                supplier_name=_clean(row.get("SNAM05")),
                entry_type=_clean(row.get("ETYP15")),
                lref=_clean(row.get("LREF15")),
                supplier_ref=_clean(row.get("SREF15")),
                po_ref=_clean(row.get("SOPN15")),
                doc_date=_ap_date(row.get("DATE"), notes, "Document date"),
                due_date=_ap_date(row.get("PDUE"), notes, "Due date"),
                period=parse_cyypp(row.get("PERIOD")),
                currency=_clean(row.get("CURN15")),
                doc_amount=doc_amount,
                open_amount=open_amount,
                gl_account=gl,
                pay_method=_clean(row.get("PMTH05")),
                pay_category=_clean(row.get("PCAT05")),
                ledger_code=_clean(row.get("LCOD15")),
                quality_notes=notes,
            ))
    return lines


# ─────────────────────────────────────────────────────────────────────
# GL listing
# ─────────────────────────────────────────────────────────────────────
@dataclass
class GLLine:
    period: str
    doc_date: datetime.date | None
    gl_account: str
    account_description: str
    flash_category: str
    channel: str
    supplier_code: str
    doc_ref: str
    po_ref: str
    description: str
    amount: float
    currency: str
    source: str
    doc_type: str
    journal_type: str
    ledger: str

    @property
    def is_ap_sourced(self) -> bool:
        """TXSRCE 'A' means the line came from the AP subledger, 'G' from a journal."""
        return self.source.upper() == "A"


def load_gl_listing(path: Path, sheet: str | None = None) -> list[GLLine]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    header_row = _find_header_row(ws, "ACCN08")
    cols = _column_map(ws, header_row)

    def get(row, name):
        idx = cols.get(name)
        return row[idx - 1] if idx and idx <= len(row) else None

    lines: list[GLLine] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        account = _clean(get(row, "ACCN08"))
        if not account:
            continue
        description = _clean(get(row, "LINDES"))
        lines.append(GLLine(
            period=parse_cyypp(get(row, "PSTPER")),
            doc_date=parse_cyymmdd(get(row, "DOCDT")),
            gl_account=account,
            account_description=_clean(get(row, "Acc Description")),
            flash_category=_clean(get(row, "Flash Category")),
            channel=_clean(get(row, "ChaNFel")) or _clean(get(row, "Channel")),
            supplier_code=_clean(get(row, "PRLACC")),
            doc_ref=_clean(get(row, "DOCREF")),
            po_ref=description if PO_PATTERN.match(description) else "",
            description=description,
            amount=_amount(get(row, "PSTAMT")),
            currency=_clean(get(row, "CURN07")),
            source=_clean(get(row, "TXSRCE")),
            doc_type=_clean(get(row, "TT")),
            journal_type=_clean(get(row, "Journal type")),
            ledger=_clean(get(row, "LEDNO")),
        ))
    return lines


# ─────────────────────────────────────────────────────────────────────
# Vendor list (VENDMAST extract)
# ─────────────────────────────────────────────────────────────────────
@dataclass
class VendorLine:
    supplier_code: str
    supplier_name: str
    company: str
    pay_method: str
    currency: str
    phone: str
    balance: float

    @property
    def is_direct_debit(self) -> bool:
        return self.pay_method.upper() == "DD"


def load_vendor_list(path: Path, sheet: str | None = None) -> list[VendorLine]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[-1]
    header_row = _find_header_row(ws, "SUPN05")
    cols = _column_map(ws, header_row)

    def get(row, name):
        idx = cols.get(name)
        return row[idx - 1] if idx and idx <= len(row) else None

    vendors: list[VendorLine] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = _clean(get(row, "SUPN05"))
        if not code:
            continue
        vendors.append(VendorLine(
            supplier_code=code,
            supplier_name=_clean(get(row, "SNAM05")),
            company=_clean(get(row, "CONO05")),
            pay_method=_clean(get(row, "PMTH05")),
            currency=_clean(get(row, "CURN05")),
            phone=_clean(get(row, "PHON05")),
            balance=_amount(get(row, "BLOG06")),
        ))
    return vendors
